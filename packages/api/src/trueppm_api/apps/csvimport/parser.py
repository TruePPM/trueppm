"""Parse a CSV / Excel upload into computable ``ProjectData`` (#743, ADR-0632).

The output is the shared ``msproject`` interchange dataclass (``ProjectData``),
so the existing, battle-tested ``msproject.importer.import_project`` persists it
— exactly as the Jira adapter does (ADR-0259). This module writes nothing; it is
Django-free apart from importing that dataclass, which is itself ORM-free.

Two shapes of hierarchy are recognized, matching what real spreadsheets carry:

- a **WBS / outline code** column ("1", "1.1", "1.2") -> dotted ``outline_number``
- **indentation** in the task-name column (leading spaces, tabs, or dots)
  -> ``outline_level``

``_build_wbs_paths`` in the importer already prefers a dotted outline number and
falls back to the level sequence, so both land on correct ltree paths without
any hierarchy code here.

Row-level errors are **data, not exceptions**: a bad date or an unresolvable
predecessor records a ``RowError`` and the row still imports with that one field
dropped. Only a structurally unusable file (no header, no name column,
unreadable bytes) raises ``CsvImportError``.

A row that cannot become a plan task at all is **parked, not dropped** (#2732):
it becomes a task under an "Import review" summary branch appended at the bottom
of the outline, carrying its raw cell values in ``notes``. Fixing one is then an
ordinary task edit — undoable and broadcast to co-authors — rather than a
re-upload of the whole file. See ``_append_review_branch``.
"""

from __future__ import annotations

import contextlib
import csv
import io
import re
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any

from trueppm_api.apps.csvimport.mapping import (
    ColumnMapping,
    detect_mapping,
    indexes_by_field,
    missing_required,
)
from trueppm_api.apps.msproject.dataclasses import (
    AssignmentData,
    PredecessorLinkData,
    ProjectData,
    ResourceData,
    TaskData,
)
from trueppm_api.core.text_decode import (
    ENCODING_ADVICE,
    TextDecodeError,
    decode_uploaded_text,
)

#: How many parsed rows the preview endpoint echoes back to the wizard.
SAMPLE_ROW_COUNT = 10

#: Ceiling on distinct labels one import may mint, and on labels per task. A
#: label column is free text under the operator's control, so an accidental
#: mapping (a Notes column landing on `labels`) would otherwise mint one catalog
#: entry per row. Both caps degrade by *dropping the excess with a warning*
#: rather than failing the import — the tasks are still worth having (#2406).
MAX_IMPORT_LABELS = 100
MAX_LABELS_PER_TASK = 20

#: `Label.name` is a CharField(max_length=50); a longer value is truncated
#: rather than dropped so the label still lands on the task.
LABEL_NAME_MAX_LENGTH = 50

#: Ceiling on inferred or declared outline depth (indentation levels, or a
#: bare-digit WBS column value). Well under Postgres's ltree label-count
#: limit -- this exists purely so a malformed or malicious cell (tens of
#: thousands of leading tabs, or a huge bare integer) can never reach
#: `_wbs_paths_from_levels` and mint an unbounded ltree path, which crashes
#: `bulk_create` with a `DataError` (#2761). A violation clamps and records a
#: row warning; the row still imports, same as every other malformed-cell
#: contract in this module.
MAX_OUTLINE_DEPTH = 100

#: Name of the summary task unresolvable rows are parked under (#2732). A fixed
#: string rather than one carrying the filename: the branch is meant to be
#: *emptied*, so the operator, the wizard copy, and the docs all have to be able
#: to name the same thing without knowing which file produced it.
REVIEW_BRANCH_NAME = "Import review"

#: Why a row could not become a plan task, phrased for the parked task's own
#: name. Keyed by the same diagnostic codes as ``SEVERITY_BY_CODE``, so a new
#: unresolvable cause declares its severity and its parked-row phrasing together.
UNRESOLVED_REASON: dict[str, str] = {
    "missing_name": "no task name",
}

#: Fallback name fragment for a parked row whose code predates the table above.
DEFAULT_UNRESOLVED_REASON = "could not be read"

#: Cap on the raw values carried onto a parked task's notes. ``Task.notes`` is an
#: unbounded TextField, but a 200-column sheet would otherwise write a wall of
#: text into the outline; the row is preserved to be *fixed*, not archived.
MAX_PARKED_VALUE_CELLS = 40

#: Per-cell cap inside that block, applied to the header **and** the value. The
#: header side is not cosmetic: header cells are parsed with no length cap, and
#: an unbounded key is copied into every parked row's notes, so a workbook with
#: 40 megabyte-long headers and 5,000 nameless rows would multiply out to
#: hundreds of gigabytes of retained strings inside the worker.
MAX_PARKED_VALUE_LENGTH = 200

#: Hard ceiling on one parked row's assembled notes, independent of the two caps
#: above so the bound survives a future change to either. ``Task.notes`` is an
#: unbounded TextField, so nothing below this enforces a limit.
MAX_PARKED_NOTES_LENGTH = 10_000

_CSV_EXTENSIONS = {"csv", "tsv", "txt"}
_XLSX_EXTENSIONS = {"xlsx", "xlsm"}
SUPPORTED_EXTENSIONS = _CSV_EXTENSIONS | _XLSX_EXTENSIONS

#: Re-exported so this module's other messages quote the one canonical sentence.
_ENCODING_ADVICE = ENCODING_ADVICE

_TRUEY = {"y", "yes", "true", "t", "1", "x", "on"}

# "3", "3FS", "3FS+2d", "5SS-1 day", "12 ff + 3"
_PREDECESSOR_RE = re.compile(
    r"^\s*(?P<ref>[^\s,;+\-]+?)\s*"
    r"(?:(?P<type>FS|SS|FF|SF)\s*)?"
    r"(?:(?P<sign>[+-])\s*(?P<lag>\d+(?:\.\d+)?)\s*(?:d|day|days)?\s*)?$",
    re.IGNORECASE,
)

#: The first numeric run in a cell, separators included so ``_parse_decimal`` can
#: decide which of them is the decimal mark. Deliberately greedy over ``.`` and
#: ``,``: the previous pattern (``-?\d+(?:\.\d+)?``) stopped at the first comma,
#: so a European "3,5" matched only "3" and a 3.5-day task imported as 3 days
#: with no warning of any kind (#2892).
_NUMBER_RE = re.compile(r"-?\d[\d.,]*")

_SLASH_DATE_RE = re.compile(r"^(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})$")

# Non-slash date layouts, tried in order against the raw cell text.
_DATE_FORMATS = (
    "%Y-%m-%d",
    "%Y/%m/%d",
    "%d %b %Y",
    "%d %B %Y",
    "%b %d %Y",
    "%B %d %Y",
    "%d-%b-%Y",
    "%d-%b-%y",
)

#: The date orders a caller may request on the import endpoints (#2926).
#:
#: ``auto`` is the shipped behavior — scan the file and settle the convention
#: from the first self-identifying value. The other three are the operator
#: asserting what their own export uses, which is the only thing that can
#: resolve a file where every value reads validly both ways.
DATE_ORDERS = ("auto", "mdy", "dmy", "iso")

#: What ``auto`` falls back to when the file identifies nothing. US English, per
#: project convention — and the reason the ambiguous state has to be announced
#: rather than assumed: it is a coin flip, not a reading.
AMBIGUOUS_FALLBACK_ORDER = "mdy"


class CsvImportError(Exception):
    """The upload is structurally unusable and no rows can be produced."""


#: Severity is **declared per diagnostic code, not decided by control flow.**
#:
#: The distinction that matters to an operator is exactly one thing: *did the row
#: land as a plan task?* A row that could not be resolved is a different event
#: from a row that imported with its start date unset, and collapsing both into
#: "7 problems" hides the only one that needs a person. Keeping the mapping as a
#: table rather than as branch structure means a reader can answer "what is fatal
#: here?" by reading one dict, and the wizard can group by severity without
#: re-deriving the rule.
#:
#: ``error``   — the row did **not** become a plan task. Since #2732 it is not
#:               dropped either: it is parked in the Import review branch with
#:               its raw values intact, so ``error_count`` counts rows awaiting
#:               a person rather than rows that vanished.
#: ``warning`` — the row imported; one field was dropped or defaulted.
SEVERITY_BY_CODE: dict[str, str] = {
    "missing_name": "error",
    "bad_date": "warning",
    "bad_duration": "warning",
    "bad_percent": "warning",
    "unknown_predecessor": "warning",
    "self_dependency": "warning",
    "finish_before_start": "warning",
    "excessive_indent_depth": "warning",
}

#: Fallback for a code added without a severity. Warning, not error: a new check
#: must not silently start dropping rows from the operator's point of view just
#: because someone forgot the table.
DEFAULT_SEVERITY = "warning"


@dataclass
class RowError:
    """One row-scoped diagnostic.

    A ``warning`` row still imports, minus the offending field; an ``error`` row
    did not import at all. Both ride back in the same list so the operator sees
    one ordered account of what happened to their file.
    """

    row: int
    column: str | None
    code: str
    message: str

    @property
    def severity(self) -> str:
        return SEVERITY_BY_CODE.get(self.code, DEFAULT_SEVERITY)

    def as_dict(self) -> dict[str, Any]:
        return {
            "row": self.row,
            "column": self.column,
            "code": self.code,
            "message": self.message,
            "severity": self.severity,
        }


@dataclass
class UnresolvedRow:
    """One data row that could not be resolved into a plan task (#2732).

    Carries the row's raw cell text verbatim, keyed by header, so the parked
    task can reproduce the spreadsheet row inside the outline. This is what makes
    "nothing is silently dropped" true rather than aspirational: an unresolvable
    row becomes a task under the Import review branch instead of leaving only a
    line number in a summary the operator may never open.
    """

    row: int
    code: str
    values: dict[str, str] = field(default_factory=dict)

    @property
    def reason(self) -> str:
        """Short phrase naming why the row could not become a plan task."""
        return UNRESOLVED_REASON.get(self.code, DEFAULT_UNRESOLVED_REASON)


@dataclass
class DateReading:
    """One convention's reading of the file — the ambiguous state's comparison row.

    An ambiguous file is one where both readings are internally valid, so the
    only honest way to present the choice is to show what each one *does* to the
    same task. ``duration_days`` is the field that makes a wrong order
    impossible to miss: 03/04 → 05/04 is either 3 days or 62.
    """

    order: str
    #: Spreadsheet row the sample was taken from, 1-based including the header.
    sample_row: int | None = None
    sample_name: str = ""
    sample_raw_start: str = ""
    start: str | None = None
    finish: str | None = None
    duration_days: int | None = None
    #: Date cells that parse under this order, and those that do not.
    values_matched: int = 0
    values_failed: int = 0
    #: Rows losing at least one date cell under this order.
    rows_unparseable: int = 0

    def as_dict(self) -> dict[str, Any]:
        return {
            "order": self.order,
            "sample_row": self.sample_row,
            "sample_name": self.sample_name,
            "sample_raw_start": self.sample_raw_start,
            "start": self.start,
            "finish": self.finish,
            "duration_days": self.duration_days,
            "values_matched": self.values_matched,
            "values_failed": self.values_failed,
            "rows_unparseable": self.rows_unparseable,
        }


@dataclass
class DateConvention:
    """How the date order was settled, and the evidence for it (#2926).

    The server has always known *why* it picked a convention and then discarded
    that reasoning into a rendered sentence. This carries the reasoning across
    the API boundary instead, because the client cannot reconstruct "row 14 is
    13/04/2026, so the file can only be day-first" from a bare enum — and a
    statement the operator cannot check against their own file is just a guess
    with better typography.
    """

    #: What the caller asked for: one of :data:`DATE_ORDERS`.
    requested: str = "auto"
    #: What will actually be used: ``mdy``/``dmy``/``iso``, never ``auto``.
    resolved: str = AMBIGUOUS_FALLBACK_ORDER
    #: What ``auto`` would have chosen. Equals ``resolved`` unless overridden —
    #: the override copy names both ("Auto would have read this as M/D/Y").
    auto_resolved: str = AMBIGUOUS_FALLBACK_ORDER
    #: True only when auto had no evidence *and* there was something to settle.
    ambiguous: bool = False
    #: ``{row, column, value, reason}`` for the value that settled it, if any.
    evidence: dict[str, Any] | None = None
    values_matched: int = 0
    values_failed: int = 0
    #: Both readings, populated only for the ambiguous case that needs them.
    readings: list[DateReading] = field(default_factory=list)
    #: False when no column is mapped to a date field — the control is inert.
    has_date_columns: bool = False

    def as_dict(self) -> dict[str, Any]:
        return {
            "requested": self.requested,
            "resolved": self.resolved,
            "auto_resolved": self.auto_resolved,
            "ambiguous": self.ambiguous,
            "evidence": self.evidence,
            "values_matched": self.values_matched,
            "values_failed": self.values_failed,
            "readings": [r.as_dict() for r in self.readings],
            "has_date_columns": self.has_date_columns,
        }


@dataclass
class ParseResult:
    """Everything both the preview and the commit path need from one upload."""

    project_data: ProjectData
    headers: list[str] = field(default_factory=list)
    mapping: list[ColumnMapping] = field(default_factory=list)
    #: First ``SAMPLE_ROW_COUNT`` data rows, raw cell text, for the wizard.
    sample_rows: list[list[str]] = field(default_factory=list)
    row_errors: list[RowError] = field(default_factory=list)
    #: Rows parked in the Import review branch rather than dropped (#2732).
    unresolved_rows: list[UnresolvedRow] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    #: Data rows read (after skipping blank and ``#``-comment rows).
    total_rows: int = 0
    #: Data rows dropped because ``max_rows`` was reached.
    truncated_rows: int = 0
    #: How the date order was settled, and the evidence for it (#2926).
    date_order: DateConvention = field(default_factory=DateConvention)
    #: Per-row date rendering for the wizard's preview table (#2926): the raw
    #: cell beside what it was read as, and the duration that produces. The
    #: duration is the column that makes a wrong order impossible to miss.
    date_preview: list[dict[str, Any]] = field(default_factory=list)

    @property
    def error_count(self) -> int:
        """Diagnostics whose row did not become a plan task."""
        return sum(1 for e in self.row_errors if e.severity == "error")

    @property
    def warning_count(self) -> int:
        """Diagnostics whose row imported with a field dropped or defaulted."""
        return sum(1 for e in self.row_errors if e.severity == "warning")

    @property
    def review_task_count(self) -> int:
        """Tasks the Import review branch adds: its summary row plus one per row.

        Zero when nothing was parked — an import with no unresolvable rows must
        not leave an empty review branch behind for someone to wonder about.
        """
        return len(self.unresolved_rows) + 1 if self.unresolved_rows else 0

    @property
    def plan_task_count(self) -> int:
        """Tasks that are the operator's actual plan, excluding the review branch.

        ``len(project_data.tasks)`` is the number of rows that will be *created*;
        this is the number that answers "how much of my plan arrived?", which is
        the one the wizard's counts and the confirm step's button are about.
        """
        return len(self.project_data.tasks) - self.review_task_count


# --- File readers --------------------------------------------------------


def _extension(filename: str) -> str:
    return filename.rsplit(".", 1)[-1].lower() if "." in filename else ""


def _decode(content: bytes) -> str:
    """Decode an uploaded delimited-text file, refusing a decode that went wrong.

    Thin translation over :func:`trueppm_api.core.text_decode.decode_uploaded_text`.
    The strategy lives there because the identical defect reached a second importer
    while this copy was private to this module (#2892 -> #2937).

    Raises:
        CsvImportError: No codec produced usable spreadsheet text.
    """
    try:
        return decode_uploaded_text(content)
    except TextDecodeError as exc:
        raise CsvImportError(str(exc)) from exc


def _read_csv_rows(content: bytes) -> tuple[list[list[str]], list[str]]:
    """Decode and split a delimited text upload into rows of raw cell text."""
    warnings: list[str] = []
    text = _decode(content)
    if not text.strip():
        raise CsvImportError("The file is empty.")

    # Sniff the delimiter from the first non-comment lines so tab- and
    # semicolon-separated exports (common from European Excel) work unchanged.
    sample = "\n".join(
        line
        for line in text.splitlines()[:20]
        if line.strip() and not line.lstrip().startswith("#")
    )
    delimiter = ","
    # A single-column file gives the sniffer nothing to work with; comma is the
    # correct fallback because it degrades to one column either way.
    with contextlib.suppress(csv.Error):
        delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    if delimiter != ",":
        warnings.append(f"Detected '{delimiter}' as the column separator.")

    rows = [list(row) for row in csv.reader(io.StringIO(text), delimiter=delimiter)]
    return rows, warnings


def _read_xlsx_rows(
    content: bytes, max_uncompressed_bytes: int
) -> tuple[list[list[str]], list[str]]:
    """Read the first worksheet of an .xlsx upload into rows of cell values.

    Guards decompression bombs before parsing. openpyxl inherits XXE hardening
    from ``defusedxml`` automatically (it is a direct dependency of this
    package), but nothing in that stack bounds the *inflated* size of a zip
    member -- a 1 MB upload can expand to gigabytes. Summing the central
    directory's declared sizes is cheap, reads no member data, and happens
    before any XML parser sees a byte.
    """
    warnings: list[str] = []
    try:
        archive = zipfile.ZipFile(io.BytesIO(content))
    except zipfile.BadZipFile as exc:
        raise CsvImportError(
            "That .xlsx file could not be opened. Re-save it from Excel and try again."
        ) from exc

    declared = sum(info.file_size for info in archive.infolist())
    if declared > max_uncompressed_bytes:
        raise CsvImportError(
            f"The workbook expands to {declared // (1024 * 1024)} MB, over the "
            f"{max_uncompressed_bytes // (1024 * 1024)} MB limit. "
            "Export the sheet as CSV instead."
        )

    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - dependency is declared
        raise CsvImportError("Excel import is unavailable on this server.") from exc

    try:
        workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise CsvImportError(
            "That .xlsx file could not be read. Re-save it from Excel and try again."
        ) from exc

    try:
        sheet_names = list(workbook.sheetnames)
        if len(sheet_names) > 1:
            warnings.append(
                f"Only the first sheet ('{sheet_names[0]}') was imported. "
                f"{len(sheet_names) - 1} other sheet(s) were ignored."
            )
        worksheet = workbook[sheet_names[0]]
        rows = [[_cell_text(cell) for cell in row] for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    return rows, warnings


def _cell_text(value: Any) -> str:
    """Render one openpyxl cell value as the text the CSV path would have seen."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


# --- Value coercion ------------------------------------------------------


def _slash_date_evidence(
    cells: list[tuple[int, str, str]],
) -> tuple[str | None, dict[str, Any] | None]:
    """Find the first value that can only be read one way, and say why.

    A slash date is self-identifying only when one component exceeds 12
    (13/04/2026 can only be D/M). We scan **every** date cell before parsing any
    of them, so one unambiguous row settles the whole file rather than letting
    consecutive rows be read under different conventions.

    Returns ``(order, evidence)`` where evidence is the ``{row, column, value,
    reason}`` the wizard quotes back at the operator. ``(None, None)`` means the
    file identifies nothing — the ambiguous case, which the caller must announce
    rather than quietly resolve.

    Args:
        cells: ``(row_number, column_header, raw_text)`` for every date cell.

    Returns:
        The settled order (``"dmy"``/``"mdy"``) and its evidence, or two Nones.
    """
    for row_number, column, raw in cells:
        match = _SLASH_DATE_RE.match(raw.strip())
        if not match:
            continue
        first, second = int(match.group(1)), int(match.group(2))
        if first > 12 and second <= 12:
            return "dmy", {
                "row": row_number,
                "column": column,
                "value": raw.strip(),
                "reason": "no_thirteenth_month",
            }
        if second > 12 and first <= 12:
            return "mdy", {
                "row": row_number,
                "column": column,
                "value": raw.strip(),
                "reason": "second_part_exceeds_twelve",
            }
    return None, None


def _parse_date(raw: str, order: str) -> date | None:
    """Read one cell under a settled date order.

    ``order`` is ``mdy``/``dmy``/``iso``, never ``auto`` — resolution happens
    once per file, not once per cell. Under ``iso`` a slash date is deliberately
    *not* reinterpreted: an operator who asserts ISO is telling us their export
    is unambiguous, so a ``03/04/2026`` in it is a malformed row to be reported,
    not a value to guess at under a convention they just ruled out.
    """
    text = (raw or "").strip()
    if not text:
        return None

    match = _SLASH_DATE_RE.match(text)
    if match:
        if order == "iso":
            return None
        a, b, year = int(match.group(1)), int(match.group(2)), int(match.group(3))
        if year < 100:
            year += 2000 if year < 70 else 1900
        day, month = (a, b) if order == "dmy" else (b, a)
        try:
            return date(year, month, day)
        except ValueError:
            return None

    # Excel sometimes hands back a full timestamp for a date-formatted cell.
    text = text.split("T")[0].split(" 00:00:00")[0]
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_decimal(text: str) -> tuple[float, bool] | None:
    """Read the first number in a cell under either decimal convention.

    Returns ``(value, had_fraction)``, or ``None`` when the cell holds no number.
    ``had_fraction`` reports whether a decimal mark was actually present, which is
    what ``_parse_percent`` needs to tell Excel's native 0..1 fraction ("0,5",
    "0.5") from an already-scaled percentage ("50").

    Why this is not a ``str.replace(",", ".")``: both marks group as well as
    separate, so the rules below resolve which role each one is playing.

    - **Both marks present** — the *last* one is the decimal mark and the other is
      grouping. This is unambiguous in every locale: ``1.234,56`` and ``1,234.56``
      both mean 1234.56.
    - **One mark, repeated** — grouping. ``1.234.567`` is not a number with two
      decimal points.
    - **One comma, exactly three trailing digits, and a leading integer part that
      is not a bare zero** — grouping. ``1,500`` reads as fifteen hundred, the US
      convention this project writes in. The leading-zero carve-out matters:
      ``0,500`` is *not* ambiguous, because no locale and no tool writes five
      hundred with a leading zero, so it can only be a decimal mark. Without the
      carve-out a German 3-decimal column turned a half-day task into a 500-day
      one and read 50 % complete as 100 % — the same silent corruption this
      function exists to end, in the opposite direction.
    - **One period, exactly three trailing digits** — decimal, so ``1.500`` stays
      1.5 exactly as it parsed before this function existed. The asymmetry with
      the comma rule is the point: it keeps every previously-correct US file
      parsing identically, so fixing the European case regresses nothing.
    - **Any other single mark** — decimal. ``3,5`` is three and a half.

    A genuinely ambiguous cell is therefore resolved by convention rather than
    left to truncate in silence, which is what ``_DURATION_RE`` used to do (#2892).
    """
    match = _NUMBER_RE.search(text)
    if not match:
        return None
    # The pattern guarantees a leading digit (after an optional sign), so the
    # rstrip below can only remove trailing separators — "5," becomes "5".
    token = match.group(0).rstrip(".,")
    negative = token.startswith("-")
    digits = token.lstrip("-")
    dots = digits.count(".")
    commas = digits.count(",")

    decimal_mark: str | None
    if dots and commas:
        decimal_mark = "." if digits.rindex(".") > digits.rindex(",") else ","
    elif dots > 1 or commas > 1:
        decimal_mark = None
    elif commas == 1:
        integer_part, fraction_part = digits.split(",")
        # A three-digit tail reads as thousands grouping — unless the integer part
        # is a bare zero, in which case grouping is impossible and it is a decimal.
        groups_thousands = len(fraction_part) == 3 and integer_part.lstrip("0") != ""
        decimal_mark = None if groups_thousands else ","
    elif dots == 1:
        decimal_mark = "."
    else:
        decimal_mark = None

    if decimal_mark is None:
        normalized = digits.replace(".", "").replace(",", "")
        had_fraction = False
    else:
        grouping = "," if decimal_mark == "." else "."
        normalized = digits.replace(grouping, "").replace(decimal_mark, ".")
        had_fraction = True

    try:
        value = float(normalized)
    except ValueError:
        return None
    return (-value if negative else value), had_fraction


def _parse_duration(raw: str) -> float | None:
    """Read a duration cell as working days ("5", "5d", "5 days", "3.5", "3,5")."""
    text = (raw or "").strip()
    if not text:
        return None
    parsed = _parse_decimal(text)
    if parsed is None:
        return None
    value, _ = parsed
    # "8h"/"16 hrs" is an effort column in hours, not days. Convert on the
    # standard 8-hour day rather than importing an 8-day task.
    if re.search(r"\b(h|hr|hrs|hour|hours)\b", text, re.IGNORECASE):
        value = value / 8.0
    if re.search(r"\b(w|wk|wks|week|weeks)\b", text, re.IGNORECASE):
        value = value * 5.0
    return value


def _parse_percent(raw: str) -> float | None:
    """Read a %-complete cell, normalizing a 0..1 fraction to 0..100."""
    text = (raw or "").strip()
    if not text:
        return None
    had_sign = "%" in text
    parsed = _parse_decimal(text)
    if parsed is None:
        return None
    value, had_fraction = parsed
    # A bare "0.5" (or "0,5") with a decimal mark and no % sign is a fraction —
    # Excel's native percent storage. "50%" and "50" are already percentages.
    # The test is `had_fraction`, not `"." in <matched text>`: keying it on the
    # literal period meant a European "0,5" could never take this branch, so a
    # half-finished task imported as 0 % complete (#2892).
    if not had_sign and 0.0 < value <= 1.0 and had_fraction:
        value *= 100.0
    return max(0.0, min(100.0, value))


def _parse_bool(raw: str) -> bool:
    return (raw or "").strip().lower() in _TRUEY


def _indent_level(raw_name: str) -> int:
    """Infer outline depth from leading whitespace or dot-leaders in a name.

    Two spaces or one tab is one level (the common Excel convention); leading
    dots ("...Design") are treated the same way, one level per dot.
    """
    leading = len(raw_name) - len(raw_name.lstrip(" \t."))
    if leading == 0:
        return 0
    prefix = raw_name[:leading]
    tabs = prefix.count("\t")
    dots = prefix.count(".")
    spaces = prefix.count(" ")
    return tabs + dots + spaces // 2


def _clamp_outline_level(
    level: int,
    result: ParseResult,
    row_number: int,
    column: str,
) -> int:
    """Cap an inferred or declared outline depth at ``MAX_OUTLINE_DEPTH``.

    Called from both hierarchy sources (name-column indentation and a
    bare-digit WBS column) right where each produces a depth, so neither can
    hand an unbounded value to the shared WBS-path builder. See
    ``MAX_OUTLINE_DEPTH`` for why the ceiling exists.
    """
    if level <= MAX_OUTLINE_DEPTH:
        return level
    result.row_errors.append(
        RowError(
            row_number,
            column,
            "excessive_indent_depth",
            f"Indent depth {level} exceeds the maximum of {MAX_OUTLINE_DEPTH}; "
            f"clamped to {MAX_OUTLINE_DEPTH}.",
        )
    )
    return MAX_OUTLINE_DEPTH


# --- Main entry point ----------------------------------------------------


def parse_spreadsheet(
    content: bytes,
    filename: str,
    *,
    column_map: dict[str, str] | None = None,
    max_rows: int = 5000,
    max_uncompressed_bytes: int = 100 * 1024 * 1024,
    project_name: str = "",
    date_order: str = "auto",
) -> ParseResult:
    """Parse an uploaded CSV/XLSX into ``ProjectData`` plus preview metadata.

    ``project_data.tasks`` holds the operator's plan **plus** the Import review
    branch, when any row could not be resolved. ``plan_task_count`` is the former
    on its own; anything reporting "how much of my plan arrived" wants that one.

    Args:
        content: Raw uploaded bytes.
        filename: Sanitized filename; used to pick a reader, and named on the
            Import review branch so a second import's parked rows are traceable
            to the file that produced them.
        column_map: Optional ``{header: field}`` overrides from the wizard's
            mapping step. Wins over auto-detection, field by field.
        max_rows: Hard cap on data rows; the remainder is reported in
            ``truncated_rows`` rather than silently dropped.
        max_uncompressed_bytes: Zip-bomb ceiling for .xlsx uploads.
        project_name: Fallback name for the ``ProjectData`` header.
        date_order: ``auto``/``mdy``/``dmy``/``iso`` (#2926). ``auto`` keeps the
            file-wide inference; the other three are the operator asserting what
            their own export uses. Date order is the last locale decision that
            was inferred and unaddressable — encoding, delimiter and decimal
            separator are all settled from the file's own evidence — and an
            inference nothing can override makes the endpoint non-deterministic
            for a scripted caller, since the same rows import differently
            depending on which values happen to disambiguate the column.

    Raises:
        CsvImportError: The file is unreadable, empty, has no header row, or has
            no column that can supply ``Task.name``.
    """
    if date_order not in DATE_ORDERS:
        raise CsvImportError(
            f"Unknown date order: {date_order}. Use one of {', '.join(DATE_ORDERS)}."
        )

    ext = _extension(filename)
    if ext in _XLSX_EXTENSIONS:
        raw_rows, warnings = _read_xlsx_rows(content, max_uncompressed_bytes)
    elif ext in _CSV_EXTENSIONS:
        raw_rows, warnings = _read_csv_rows(content)
    else:
        raise CsvImportError(f"Unsupported file type: .{ext}. Upload a .csv or .xlsx file.")

    rows = [r for r in raw_rows if not _is_skippable(r)]
    if not rows:
        raise CsvImportError("The file has no readable rows.")

    headers = [str(cell).strip() for cell in rows[0]]
    if not any(headers):
        raise CsvImportError("The first row must be a header row naming each column.")

    mapping = detect_mapping(headers, overrides=column_map)
    missing = missing_required(mapping)
    if missing:
        raise CsvImportError(
            "No column could be matched to the task name. "
            "Rename a column to 'Name' or 'Task', or map one explicitly."
        )

    by_field = indexes_by_field(mapping)
    data_rows = rows[1:]

    truncated = 0
    if len(data_rows) > max_rows:
        truncated = len(data_rows) - max_rows
        warnings.append(
            f"Only the first {max_rows:,} rows were imported; {truncated:,} were skipped."
        )
        data_rows = data_rows[:max_rows]

    result = ParseResult(
        project_data=ProjectData(name=project_name[:255]),
        headers=headers,
        mapping=mapping,
        row_errors=[],
        warnings=warnings,
        total_rows=len(data_rows),
        truncated_rows=truncated,
    )
    result.sample_rows = [
        [_cell(row, i) for i in range(len(headers))] for row in data_rows[:SAMPLE_ROW_COUNT]
    ]

    _build_tasks(result, data_rows, by_field, date_order)
    _resolve_predecessors(result, data_rows, by_field)
    _build_resources(result, data_rows, by_field)
    _build_labels(result, data_rows, by_field)
    # Last, and deliberately so — see _append_review_branch.
    _append_review_branch(result, filename)

    result.project_data.warnings = list(result.warnings)
    return result


def _is_skippable(row: list[Any]) -> bool:
    """True for a blank row or a ``#``-prefixed comment row."""
    cells = [str(c).strip() for c in row if c is not None]
    if not any(cells):
        return True
    return bool(cells) and cells[0].startswith("#")


def _cell(row: list[Any], index: int | None) -> str:
    if index is None or index >= len(row) or row[index] is None:
        return ""
    return str(row[index])


def _one(by_field: dict[str, list[int]], field: str) -> int:
    """The single column supplying ``field``.

    ``by_field`` is a list per field so that a ``multi`` field's extra columns
    cannot be silently dropped by a scalar lookup. Every single-valued read goes
    through here, which makes "this field is one-to-one" an explicit claim at
    the call site rather than an assumption baked into the container's shape.

    Only call for a field known to be present — callers gate on
    ``field in by_field`` exactly as they did when this was a scalar map.
    """
    return by_field[field][0]


def _split_values(raw: str) -> list[str]:
    """Split one cell into its values on the shared ``, ; /`` convention.

    Same separators the resource column has always used, so a sheet author who
    learned the convention in one place does not have to relearn it.
    """
    return [part.strip() for part in re.split(r"[,;/]", raw) if part.strip()]


def _wbs_column_holds_codes(data_rows: list[list[Any]], by_field: dict[str, list[int]]) -> bool:
    """Decide once, for the whole column, whether its cells are outline codes or depths.

    Whether a WBS column holds *codes* (``1``, ``1.1``, ``2``) or *depths*
    (``0``, ``1``, ``1``) is a property of the column, not of the individual cell — but
    ``_apply_wbs`` used to decide per row, on whether that one value contained a dot.
    A real-world WBS column mixes both spellings, because a phase's own code is bare and
    only its children are dotted. Every dotted row was then read as a code and every
    phase row as a depth, inside one file (#3082).

    The phase rows are the ones that break. A bare cell taken as a depth leaves
    ``outline_number`` at the positional fallback ``_build_task`` assigned, so the second
    phase of a file lands at its row index instead of its code — and since ``wbs_path``
    is the only record of parenthood, its children are left pointing at a number no live
    row holds. They render at project root, and the WBS has a hole where the phase was.

    The first phase always looks correct, which is why this went unnoticed: it is row 1
    and its code is ``1``, so the fallback and the intended code agree by coincidence.

    One dotted cell anywhere settles the file, for the same reason one unambiguous slash
    date does in :func:`_slash_date_evidence`: a convention that consecutive rows are
    allowed to disagree about is not a convention. A column with no dotted cell at all is
    the ``Level``/``Outline`` depth convention ``_apply_wbs`` was written for, and keeps
    being read that way.
    """
    if "wbs" not in by_field:
        return False
    wbs_index = _one(by_field, "wbs")
    return any("." in _cell(row, wbs_index) for row in data_rows)


def _build_tasks(
    result: ParseResult,
    data_rows: list[list[Any]],
    by_field: dict[str, list[int]],
    date_order: str,
) -> None:
    """Turn each data row into a ``TaskData``, recording per-field failures."""
    name_index = _one(by_field, "name")
    result.date_order = _resolve_date_convention(result, data_rows, by_field, date_order)
    order = result.date_order.resolved
    # Settled across every row before any is read, not per cell — see the function's
    # docstring for why the per-cell call orphaned every phase but the first (#3082).
    wbs_is_codes = _wbs_column_holds_codes(data_rows, by_field)
    tasks: list[TaskData] = []

    for offset, row in enumerate(data_rows):
        # Row numbers are 1-based and count the header, so they line up with
        # what the operator sees in Excel's row gutter.
        row_number = offset + 2
        raw_name = _cell(row, name_index)
        if not raw_name.strip():
            result.row_errors.append(
                RowError(
                    row_number,
                    result.headers[name_index],
                    "missing_name",
                    "Row has no task name; parked in the Import review branch.",
                )
            )
            result.unresolved_rows.append(
                UnresolvedRow(row_number, "missing_name", _raw_values(result.headers, row))
            )
            continue
        tasks.append(
            _build_task(
                offset,
                raw_name,
                row,
                by_field,
                result,
                row_number,
                order,
                len(tasks),
                wbs_is_codes=wbs_is_codes,
            )
        )

    result.project_data.tasks = tasks
    _build_date_preview(result, data_rows, by_field, order)


def _build_date_preview(
    result: ParseResult,
    data_rows: list[list[Any]],
    by_field: dict[str, list[int]],
    order: str,
) -> None:
    """Render the first rows' dates as the wizard's preview table shows them.

    Built server-side beside the parse rather than re-derived in the client,
    for the same reason the mapping is: the operator is confirming what the
    parser will actually do, and a second implementation of "read this cell as
    a date" is a second thing that can disagree with the import.
    """
    if not result.date_order.has_date_columns:
        return
    start_idx = _one(by_field, "planned_start") if "planned_start" in by_field else None
    finish_idx = _one(by_field, "planned_finish") if "planned_finish" in by_field else None
    name_idx = _one(by_field, "name") if "name" in by_field else None

    rows: list[dict[str, Any]] = []
    for offset, row in enumerate(data_rows[:SAMPLE_ROW_COUNT]):
        raw_start = _cell(row, start_idx).strip() if start_idx is not None else ""
        raw_finish = _cell(row, finish_idx).strip() if finish_idx is not None else ""
        start = _parse_date(raw_start, order) if raw_start else None
        finish = _parse_date(raw_finish, order) if raw_finish else None
        rows.append(
            {
                "row": offset + 2,
                "name": _cell(row, name_idx).strip() if name_idx is not None else "",
                "raw_start": raw_start,
                "raw_finish": raw_finish,
                "start": start.isoformat() if start else None,
                "finish": finish.isoformat() if finish else None,
                # Inclusive of both endpoints, matching _apply_derived_duration.
                "duration_days": (finish - start).days + 1 if start and finish else None,
                # A cell that was present and would not parse — the state the
                # "{k} of {n} rows cannot be read as {order}" copy counts.
                "unreadable": bool((raw_start and not start) or (raw_finish and not finish)),
            }
        )
    result.date_preview = rows


# --- Import review branch (#2732) ----------------------------------------


def _raw_values(headers: list[str], row: list[Any]) -> dict[str, str]:
    """Snapshot one spreadsheet row's non-empty cells as ``{header: text}``.

    Taken before any coercion so the parked task shows what the operator wrote,
    not what the parser made of it — the point of parking a row is that the
    parser's reading of it is the thing in question.

    The **key** is truncated as well as the value. Header cells carry no length
    cap of their own, and this dict is materialized once per parked row, so an
    uncapped key is the one term here that multiplies: a workbook of huge headers
    plus thousands of nameless rows would otherwise expand a sub-megabyte upload
    into hundreds of gigabytes of strings in the worker.
    """
    values: dict[str, str] = {}
    for index, header in enumerate(headers[:MAX_PARKED_VALUE_CELLS]):
        text = _cell(row, index).strip()
        if text:
            key = (header or f"Column {index + 1}")[:MAX_PARKED_VALUE_LENGTH]
            values[key] = text[:MAX_PARKED_VALUE_LENGTH]
    return values


def _next_top_level_number(tasks: list[TaskData]) -> int:
    """The first top-level outline number no imported task has claimed.

    Read from the *first* segment of each outline number so a file using dotted
    WBS codes ("1", "1.1", "2.3") yields the next whole-numbered phase. The
    length gate mirrors ``_apply_wbs``: an absurdly long digit string is a
    malformed cell, not a real outline position, and must not reach ``int()``.
    """
    highest = 0
    for task in tasks:
        head = (task.outline_number or "").split(".")[0].strip()
        if head.isdigit() and len(head) <= 10:
            highest = max(highest, int(head))
    return highest + 1


def _append_review_branch(result: ParseResult, filename: str) -> None:
    """Park every unresolvable row under an "Import review" summary task.

    An unresolvable row used to be counted and then dropped, which meant the
    operator's only record of it was a line number in a summary they had to be
    looking at when the import finished. Parking the row keeps the data in the
    plan, where a normal task edit fixes it — undoable, and broadcast to
    co-authors like any other edit, with no special repair surface to build.

    **Appended last, after predecessor / resource / label resolution.** Those
    passes address ``data_rows`` by ``task.uid``, so a synthetic uid in the task
    list would index out of range. Running after them also means the parked rows
    carry no predecessor links, so the dependency graph the caller hands to
    ``validate_task_graph`` (ADR-0259) is exactly the one the real rows produced.

    Hierarchy is expressed in whichever of ``_build_wbs_paths``' two modes the
    real rows already use. Writing a dotted outline number onto a file whose
    hierarchy lives in indentation would flip that function's ``has_dotted``
    branch for the *whole* import and re-derive every real task's path from a
    number the parser only ever meant as a sequence — so the dotted form is used
    only when the file was already dotted, and levels carry it otherwise.

    ``_build_wbs_paths`` has a **second** discriminator, though, and this branch
    unavoidably trips it: a genuinely flat file has ``len(set(levels)) <= 1`` and
    derives paths from the outline number, and adding a summary plus a child
    makes the level set non-trivial, so it switches to the level sequence. That
    is safe here only because of an invariant elsewhere in this module — when the
    WBS column is not dotted, ``_apply_wbs`` reads a bare integer as a *depth*
    and leaves ``outline_number`` as the sequence ``_build_task`` assigned, so
    both modes produce the same 1..N paths for the real rows. If ``_apply_wbs``
    ever preserves a non-sequential flat code, one parked row would silently
    renumber the whole import; a test pins the indentation case.
    """
    if not result.unresolved_rows:
        return

    tasks = result.project_data.tasks
    has_dotted = any("." in (t.outline_number or "") for t in tasks)
    base_level = min((t.outline_level for t in tasks), default=0)
    # Only meaningful in dotted mode; left empty otherwise rather than written
    # and ignored, so the field never disagrees with the path actually derived.
    next_top = str(_next_top_level_number(tasks)) if has_dotted else ""
    next_uid = max((t.uid for t in tasks), default=0) + 1

    tasks.append(
        TaskData(
            uid=next_uid,
            name=REVIEW_BRANCH_NAME,
            outline_number=next_top,
            outline_level=base_level,
            notes=(
                f"Rows from {filename} that could not be imported as tasks. "
                "Each one below keeps the values you wrote. Fix or delete them, "
                "and this branch can go."
            ),
        )
    )
    for offset, unresolved in enumerate(result.unresolved_rows, start=1):
        tasks.append(
            TaskData(
                uid=next_uid + offset,
                name=f"Row {unresolved.row} — {unresolved.reason}"[:255],
                # Dotted only when the file was: see the docstring.
                outline_number=f"{next_top}.{offset}" if next_top else "",
                outline_level=base_level + 1,
                notes=_parked_notes(unresolved),
            )
        )


def _parked_notes(unresolved: UnresolvedRow) -> str:
    """Render one parked row's raw values as the task's notes.

    ``Header: value`` lines rather than the original delimited text: the row is
    here to be read and fixed by a person in the outline, and a re-serialized CSV
    line would make them count commas to find the cell that is wrong.

    Truncated at ``MAX_PARKED_NOTES_LENGTH`` as a backstop. The per-cell caps
    already bound this, but ``Task.notes`` is an unbounded TextField and this
    string is written once per parked row, so the ceiling is stated here too
    rather than left as a consequence of two other constants.
    """
    lines = [
        f"Spreadsheet row {unresolved.row} — {unresolved.reason}.",
        "",
    ]
    if unresolved.values:
        lines += [f"{header}: {value}" for header, value in unresolved.values.items()]
    else:
        lines.append("(the row was empty in every mapped column)")
    return "\n".join(lines)[:MAX_PARKED_NOTES_LENGTH]


def _resolve_date_convention(
    result: ParseResult,
    data_rows: list[list[Any]],
    by_field: dict[str, list[int]],
    requested: str,
) -> DateConvention:
    """Settle the date order once, over every date cell, and record the evidence.

    Resolved file-wide rather than per row so that a single unambiguous value
    (13/04 can only be day-first) settles the whole import — a per-row guess
    would read consecutive rows under different conventions.

    An explicit ``requested`` order always wins: the operator knows what their
    own export contains, and a heuristic that overrode them would make the
    control decorative. ``auto`` keeps the shipped behavior, and additionally
    reports *why* it chose what it chose, plus both readings when it could not
    tell (which is the state this whole feature exists for — see #2926).

    Args:
        result: Parse result; its ``headers`` name the columns in the evidence.
        data_rows: Every data row, so one late unambiguous value still counts.
        by_field: Column indexes per mapped field.
        requested: One of :data:`DATE_ORDERS`.

    Returns:
        The convention, its evidence, and the per-order tallies the wizard needs.
    """
    date_indices = [_one(by_field, f) for f in ("planned_start", "planned_finish") if f in by_field]
    convention = DateConvention(requested=requested, has_date_columns=bool(date_indices))
    if not date_indices:
        # Nothing is mapped to a date field, so there is no convention to settle
        # and nothing for the control to act on. Reported rather than guessed:
        # the wizard renders an inert block instead of a confident-looking one.
        return convention

    cells = [
        (offset + 2, result.headers[i], _cell(row, i))
        for offset, row in enumerate(data_rows)
        for i in date_indices
        if _cell(row, i).strip()
    ]
    has_slash = any(_SLASH_DATE_RE.match(raw.strip()) for _, _, raw in cells)

    auto_order, evidence = _slash_date_evidence(cells)
    if auto_order is None and not has_slash and cells:
        # Every value is a non-slash layout, so the file is self-describing and
        # there was never a decision to make. Named as ISO because that is what
        # the operator sees offered, and what the vast majority of such files are.
        auto_order, evidence = (
            "iso",
            {
                "row": cells[0][0],
                "column": cells[0][1],
                "value": cells[0][2].strip(),
                "reason": "non_slash_layout",
            },
        )
    convention.auto_resolved = auto_order or AMBIGUOUS_FALLBACK_ORDER

    if requested == "auto":
        convention.resolved = convention.auto_resolved
        convention.evidence = evidence
        # Ambiguous only when auto had a real decision to make and no evidence
        # for it. A file with no slash dates at all is not ambiguous; it is
        # simply unambiguous in a way that needed no argument.
        convention.ambiguous = auto_order is None and has_slash
    else:
        convention.resolved = requested
        convention.evidence = evidence
        convention.ambiguous = False

    convention.values_matched, convention.values_failed = _tally(cells, convention.resolved)

    if convention.ambiguous:
        convention.readings = [
            _reading(order, cells, data_rows, by_field, result) for order in ("mdy", "dmy")
        ]
    return convention


def _tally(cells: list[tuple[int, str, str]], order: str) -> tuple[int, int]:
    """Date cells that parse under ``order``, and those that do not."""
    matched = sum(1 for _, _, raw in cells if _parse_date(raw, order) is not None)
    return matched, len(cells) - matched


def _reading(
    order: str,
    cells: list[tuple[int, str, str]],
    data_rows: list[list[Any]],
    by_field: dict[str, list[int]],
    result: ParseResult,
) -> DateReading:
    """What one convention does to this file — tallies plus one worked example.

    The sample is the first row that has both a start and a finish under this
    order, because a duration is the only rendering of the choice that is
    self-evidently right or wrong to a reader.
    """
    matched, failed = _tally(cells, order)
    reading = DateReading(order=order, values_matched=matched, values_failed=failed)

    start_idx = _one(by_field, "planned_start") if "planned_start" in by_field else None
    finish_idx = _one(by_field, "planned_finish") if "planned_finish" in by_field else None
    name_idx = _one(by_field, "name") if "name" in by_field else None

    for offset, row in enumerate(data_rows):
        raws = [_cell(row, i) for i in (start_idx, finish_idx) if i is not None]
        if any(raw.strip() and _parse_date(raw, order) is None for raw in raws):
            reading.rows_unparseable += 1
        if reading.sample_row is not None:
            continue
        start = _parse_date(_cell(row, start_idx), order) if start_idx is not None else None
        finish = _parse_date(_cell(row, finish_idx), order) if finish_idx is not None else None
        if start and finish:
            reading.sample_row = offset + 2
            reading.sample_name = _cell(row, name_idx).strip() if name_idx is not None else ""
            reading.sample_raw_start = _cell(row, start_idx).strip()
            reading.start = start.isoformat()
            reading.finish = finish.isoformat()
            # Inclusive of both endpoints, matching _apply_derived_duration —
            # the comparison is worthless if it disagrees with the duration the
            # import will actually write.
            reading.duration_days = (finish - start).days + 1
    return reading


def _build_task(
    offset: int,
    raw_name: str,
    row: list[Any],
    by_field: dict[str, list[int]],
    result: ParseResult,
    row_number: int,
    order: str,
    task_count: int,
    *,
    wbs_is_codes: bool,
) -> TaskData:
    """Build one ``TaskData`` from a named row, applying every mapped column."""
    task = TaskData(uid=offset + 1, name=raw_name.strip()[:255])
    task.outline_number = str(task_count + 1)
    task.outline_level = _clamp_outline_level(
        _indent_level(raw_name), result, row_number, result.headers[_one(by_field, "name")]
    )

    if "wbs" in by_field:
        wbs_index = _one(by_field, "wbs")
        wbs_raw = _cell(row, wbs_index).strip()
        if wbs_raw:
            _apply_wbs(
                task,
                wbs_raw,
                result,
                row_number,
                result.headers[wbs_index],
                is_code=wbs_is_codes,
            )

    _apply_duration(task, row, by_field, result, row_number)
    _apply_dates(task, row, by_field, result, row_number, order)
    _apply_percent(task, row, by_field, result, row_number)

    if "milestone" in by_field and _parse_bool(_cell(row, _one(by_field, "milestone"))):
        task.is_milestone = True
    # A zero-day task is a milestone whatever the Milestone column says — the
    # duration is the stronger signal because it is what the schedule computes on.
    if task.duration_days == 0:
        task.is_milestone = True
    if "notes" in by_field:
        task.notes = _cell(row, _one(by_field, "notes")).strip()[:2000]
    return task


def _apply_wbs(
    task: TaskData,
    wbs_raw: str,
    result: ParseResult,
    row_number: int,
    column: str,
    *,
    is_code: bool,
) -> None:
    """Apply one WBS cell under the convention settled for its whole column.

    ``is_code`` comes from :func:`_wbs_column_holds_codes`, which reads every cell in
    the column before any of them is applied. It is not re-derived here: deciding per
    cell is the defect this argument exists to remove (#3082), and a bare integer is a
    perfectly ordinary *code* in a column whose other rows are dotted.
    """
    if is_code and all(p.strip().isdigit() for p in wbs_raw.split(".")):
        # An outline code carries the hierarchy directly; the importer's
        # _build_wbs_paths prefers this over the level sequence. A bare integer here is
        # a top-level code (a phase's own number), not a depth.
        task.outline_number = wbs_raw.strip()
        task.outline_level = wbs_raw.count(".")
    elif not is_code and wbs_raw.isdigit():
        # A bare integer in a "Level"/"Outline" column is a depth, not a code.
        # The length gate avoids handing an enormous digit string straight to
        # int(): Python's int-from-str conversion limit would raise before
        # _clamp_outline_level ever runs, and no legitimate depth is remotely
        # this long (#2761).
        level = int(wbs_raw) if len(wbs_raw) <= 10 else MAX_OUTLINE_DEPTH + 1
        task.outline_level = _clamp_outline_level(level, result, row_number, column)


def _apply_duration(
    task: TaskData,
    row: list[Any],
    by_field: dict[str, list[int]],
    result: ParseResult,
    row_number: int,
) -> None:
    if "duration" not in by_field:
        return
    raw = _cell(row, _one(by_field, "duration"))
    if not raw.strip():
        return
    value = _parse_duration(raw)
    if value is None:
        result.row_errors.append(
            RowError(
                row_number,
                result.headers[_one(by_field, "duration")],
                "bad_duration",
                f"Could not read '{raw.strip()}' as a duration; defaulted to 1 day.",
            )
        )
        return
    if value < 0:
        result.row_errors.append(
            RowError(
                row_number,
                result.headers[_one(by_field, "duration")],
                "bad_duration",
                f"Negative duration '{raw.strip()}'; defaulted to 1 day.",
            )
        )
        return
    task.duration_days = round(value)


def _apply_dates(
    task: TaskData,
    row: list[Any],
    by_field: dict[str, list[int]],
    result: ParseResult,
    row_number: int,
    order: str,
) -> None:
    start = _read_date_cell("planned_start", "start", row, by_field, result, row_number, order)
    finish = _read_date_cell("planned_finish", "finish", row, by_field, result, row_number, order)

    if start:
        task.start = start.isoformat()

    # ProjectData carries a start + duration, not a finish, so a file that gives
    # both has its duration derived from the span (inclusive of both endpoints).
    # An explicit Duration column already read above wins, because it is the
    # operator's stated intent rather than something inferred.
    if start and finish and "duration" not in by_field:
        _apply_derived_duration(task, start, finish, by_field, result, row_number)


def _read_date_cell(
    field_name: str,
    what: str,
    row: list[Any],
    by_field: dict[str, list[int]],
    result: ParseResult,
    row_number: int,
    order: str,
) -> date | None:
    """Read one date column, recording a row error when the cell will not parse.

    An unreadable date is data, not an exception: the row still imports with that
    one field dropped, so a single malformed cell cannot cost the operator the
    whole file.
    """
    if field_name not in by_field:
        return None
    raw = _cell(row, _one(by_field, field_name))
    if not raw.strip():
        return None
    value = _parse_date(raw, order)
    if value is None:
        result.row_errors.append(
            RowError(
                row_number,
                result.headers[_one(by_field, field_name)],
                "bad_date",
                f"Could not read '{raw.strip()}' as a date; the {what} was left unset.",
            )
        )
    return value


def _apply_derived_duration(
    task: TaskData,
    start: date,
    finish: date,
    by_field: dict[str, list[int]],
    result: ParseResult,
    row_number: int,
) -> None:
    """Set duration from an inclusive ``start``..``finish`` span."""
    span = (finish - start).days + 1
    if span < 1:
        result.row_errors.append(
            RowError(
                row_number,
                result.headers[_one(by_field, "planned_finish")],
                "finish_before_start",
                "Finish is before start; duration defaulted to 1 day.",
            )
        )
    else:
        task.duration_days = span


def _apply_percent(
    task: TaskData,
    row: list[Any],
    by_field: dict[str, list[int]],
    result: ParseResult,
    row_number: int,
) -> None:
    if "percent_complete" not in by_field:
        return
    raw = _cell(row, _one(by_field, "percent_complete"))
    if not raw.strip():
        return
    value = _parse_percent(raw)
    if value is None:
        result.row_errors.append(
            RowError(
                row_number,
                result.headers[_one(by_field, "percent_complete")],
                "bad_percent",
                f"Could not read '{raw.strip()}' as a percentage; treated as 0%.",
            )
        )
        return
    task.percent_complete = value


def _resolve_predecessors(
    result: ParseResult,
    data_rows: list[list[Any]],
    by_field: dict[str, list[int]],
) -> None:
    """Attach FS/SS/FF/SF links, quarantining refs that do not resolve.

    References are matched against the source file's own ID column when it has
    one (how MS Project and Asana CSV exports encode dependencies), and against
    1-based row position otherwise. Both are tried for every reference, so a
    file that mixes the two still resolves.

    ``predecessors`` is a ``multi`` field (#2406): an MS Project export spreads
    dependencies across ``Predecessor 1``, ``Predecessor 2``, … and every one of
    them is a real link. Each mapped column is read, and a row error names the
    column the bad token actually came from rather than the first one.
    """
    indexes = by_field.get("predecessors", [])
    if not indexes:
        return

    tasks = result.project_data.tasks
    by_handle = _predecessor_handle_map(tasks, data_rows, by_field)
    for task in tasks:
        _attach_row_predecessors(task, indexes, by_handle, data_rows, result)


def _predecessor_handle_map(
    tasks: list[TaskData], data_rows: list[list[Any]], by_field: dict[str, list[int]]
) -> dict[str, int]:
    """Map every usable predecessor handle → task uid.

    Row position is registered first so an explicit ID column overrides it on
    collision — the ID is the stronger claim.
    """
    by_handle: dict[str, int] = {
        str(position): task.uid for position, task in enumerate(tasks, start=1)
    }
    if "external_id" not in by_field:
        return by_handle

    id_index = _one(by_field, "external_id")
    # data_rows and tasks diverge when a row was dropped for a missing name,
    # so walk the tasks and re-read the row each one came from via its uid.
    for task in tasks:
        raw_id = _cell(data_rows[task.uid - 1], id_index).strip()
        if raw_id:
            by_handle[raw_id] = task.uid
    return by_handle


def _resolved_link(
    token: str,
    by_handle: dict[str, int],
    task: TaskData,
    *,
    row_number: int,
    column: str,
    result: ParseResult,
) -> PredecessorLinkData | None:
    """Resolve one reference token, or record why it was skipped and return ``None``.

    A bad token never fails the import — it quarantines as a row error naming the
    column it actually came from, and the rest of the row still links.
    """
    link = _parse_predecessor(token, by_handle)
    if link is None:
        result.row_errors.append(
            RowError(
                row_number,
                column,
                "unknown_predecessor",
                f"Predecessor '{token.strip()}' does not match any row; the link was skipped.",
            )
        )
        return None
    if link.predecessor_uid == task.uid:
        result.row_errors.append(
            RowError(
                row_number,
                column,
                "self_dependency",
                "A task cannot depend on itself; the link was skipped.",
            )
        )
        return None
    return link


def _attach_row_predecessors(
    task: TaskData,
    indexes: list[int],
    by_handle: dict[str, int],
    data_rows: list[list[Any]],
    result: ParseResult,
) -> None:
    """Read every mapped predecessor column for one task and attach its links."""
    row_number = task.uid + 1
    seen: set[tuple[int, str, int]] = set()
    for index in indexes:
        column = result.headers[index]
        raw = _cell(data_rows[task.uid - 1], index).strip()
        if not raw:
            continue
        for token in re.split(r"[,;]", raw):
            if not token.strip():
                continue
            link = _resolved_link(
                token, by_handle, task, row_number=row_number, column=column, result=result
            )
            if link is None:
                continue
            # Two columns naming the same predecessor is a union, not two
            # edges — the importer would otherwise create a duplicate
            # Dependency row for one real relationship.
            key = (link.predecessor_uid, link.dep_type, link.lag_days)
            if key in seen:
                continue
            seen.add(key)
            task.predecessor_links.append(link)


def _parse_predecessor(token: str, by_handle: dict[str, int]) -> PredecessorLinkData | None:
    match = _PREDECESSOR_RE.match(token)
    if not match:
        return None
    uid = by_handle.get(match.group("ref").strip())
    if uid is None:
        return None
    lag = float(match.group("lag") or 0)
    if match.group("sign") == "-":
        lag = -lag
    return PredecessorLinkData(
        predecessor_uid=uid,
        dep_type=(match.group("type") or "FS").upper(),
        lag_days=round(lag),
    )


def _build_labels(
    result: ParseResult,
    data_rows: list[list[Any]],
    by_field: dict[str, list[int]],
) -> None:
    """Attach label names to each task, unioned across every mapped column.

    ``labels`` is a ``multi`` field (#2406): a sheet routinely carries `Tags`,
    `Component` and `Team`, all of which are labels. Every mapped column is
    read and the values unioned onto the task.

    Names are deduplicated case-insensitively — first-seen spelling wins, both
    within a task and across the file — so an import cannot mint `Safety`
    alongside an existing `safety`. The parser only produces *names*; matching
    them against the project's existing catalog is the importer's job, since
    only it can see the database.
    """
    indexes = by_field.get("labels", [])
    if not indexes:
        return

    # File-wide canonical spelling per lowercased name, so two tasks spelling a
    # label differently still land on one catalog entry.
    canonical: dict[str, str] = {}
    truncated = False

    for task in result.project_data.tasks:
        dropped, row_truncated = _collect_task_labels(
            task, data_rows[task.uid - 1], indexes, canonical
        )
        truncated = truncated or row_truncated
        if dropped:
            result.row_errors.append(
                RowError(
                    task.uid + 1,
                    result.headers[indexes[0]],
                    "too_many_labels",
                    f"{dropped} label(s) on this row were skipped; an import is capped at "
                    f"{MAX_IMPORT_LABELS} distinct labels and {MAX_LABELS_PER_TASK} per task.",
                )
            )

    if truncated:
        result.warnings.append(
            f"Some label names were longer than {LABEL_NAME_MAX_LENGTH} characters "
            "and were shortened."
        )


def _collect_task_labels(
    task: TaskData,
    row: list[Any],
    indexes: list[int],
    canonical: dict[str, str],
) -> tuple[int, bool]:
    """Union every mapped label column onto one task, honoring the import caps.

    ``canonical`` is threaded in and mutated so the first-seen spelling of a name
    wins file-wide, not merely within this row.

    Returns ``(dropped, truncated)`` — how many names the caps skipped, and
    whether any name had to be shortened.
    """
    seen: set[str] = set()
    dropped = 0
    truncated = False
    for index in indexes:
        for raw_name in _split_values(_cell(row, index)):
            name = raw_name[:LABEL_NAME_MAX_LENGTH]
            truncated = truncated or len(raw_name) > LABEL_NAME_MAX_LENGTH
            key = name.lower()
            if key in seen:
                continue
            if key not in canonical and len(canonical) >= MAX_IMPORT_LABELS:
                dropped += 1
                continue
            if len(seen) >= MAX_LABELS_PER_TASK:
                dropped += 1
                continue
            seen.add(key)
            task.labels.append(canonical.setdefault(key, name))
    return dropped, truncated


def _build_resources(
    result: ParseResult,
    data_rows: list[list[Any]],
    by_field: dict[str, list[int]],
) -> None:
    """Collect distinct assignee names and attach one assignment per task.

    Names are deduplicated case-insensitively but the first-seen spelling is
    kept, so "J. Smith" and "j. smith" become one resource named as first
    written. ``import_project`` match-or-creates these against the project's
    resource pool.
    """
    if "resource" not in by_field:
        return

    resources: list[ResourceData] = []
    uid_by_key: dict[str, int] = {}

    for task in result.project_data.tasks:
        raw = _cell(data_rows[task.uid - 1], _one(by_field, "resource")).strip()
        if not raw:
            continue
        for part in re.split(r"[,;/]", raw):
            name = part.strip()
            if not name:
                continue
            key = name.lower()
            if key not in uid_by_key:
                uid_by_key[key] = len(resources) + 1
                resources.append(ResourceData(uid=uid_by_key[key], name=name[:255]))
            task.resource_assignments.append(
                AssignmentData(task_uid=task.uid, resource_uid=uid_by_key[key], units=1.0)
            )

    result.project_data.resources = resources
